import sys

from openpyxl import load_workbook

def get_workbook(file):
    workbook = load_workbook(file, keep_vba=True, data_only=True)
    a = workbook.sheetnames
    ws = workbook[a[0]]
    M = {}
    for row in ws.rows:
        if not(M):
            M = {i.value:[] for i in row}
        else:
            b = ["".join(str(i.value).splitlines()) for i in row]
            for keys in M:
                M[keys].append(str(b[0]))
                del b[0]
    return M
current_state = 0
tree_judgment_all = get_workbook("ts_6/med.xlsx")
rules_all = get_workbook("ts_6/rule.xlsx")
disease_by_id = get_workbook("ts_6/disease.xlsx")


def get_judg(current_state):
    return [tree_judgment_all[keys][tree_judgment_all["condition"].index(str(current_state))] for keys in tree_judgment_all]

def disease_bad(current_state):
    items = get_judg(current_state)
    print(items)
    if (len(items) == 0):
        sys.exit('ошибка, невозможное осстояние')
    return items[4] == str(-1)

def get_rules(current_state):
    indexes = [i for i in range(len(rules_all["condition"])) if rules_all["condition"][i] == str(current_state)]
    return [[rules_all[keys][i] for keys in rules_all] for i in indexes]

def get_answer(rules):
    ans = int(input())
    if (ans <= len(rules_all)):
        return ans

def get_disease_by_id(id):
    if id in disease_by_id["id"]:
        return [disease_by_id[keys][disease_by_id["id"].index(str(id))] for keys in disease_by_id]
    return []

while (disease_bad(current_state)):
    tree_judgment = get_judg(current_state)
    print('Выберите один из вариантов ответа')
    rules = get_rules(current_state)
    for i in range(len(rules)):
        print(str(i + 1) + ')', rules[i][2])
    answer = get_answer(rules)
    current_state = rules[answer - 1][3]


tree_judgment = get_judg(current_state)
disease = get_disease_by_id(tree_judgment[4])
if (not(disease)):
    print('Ваще заболевание не определенно')
elif (len(disease)):
    print('Возможно у вас:')
    print(disease[0])
    print(disease[2])
    print('Предлагаем вам лечение:')
    print(disease[4])
    print('Обратитесь к врачу')
